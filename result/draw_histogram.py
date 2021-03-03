# -*-coding:UTF-8 -*-
'''
* draw_histogram.py
* @author wzm
* created 2021/03/03 14:24:59
* @function: 绘制对比柱状图
'''

import xlrd
import openpyxl

import matplotlib
import matplotlib.pyplot as plt
import seaborn as sns
sns.set(color_codes=True)
matplotlib.rcParams["font.sans-serif"] = ["SimHei"]
matplotlib.rcParams["axes.unicode_minus"] = False

import numpy as np

import pandas as pd

def get_excel_data(excel_path):
    excel_sheets = openpyxl.load_workbook(excel_path)
    
    accuracy_sheet = excel_sheets['accuracy']
    print("行数", accuracy_sheet.max_row)
    print("列数", accuracy_sheet.max_column)
    
    label_list = []
    xlabel_list = []
    origin_list, pretrained_list, spp_list = [], [], []
    
    for i, row in enumerate(accuracy_sheet.rows):
        if i == 0: # 将第一行的数据作为标签
            for cell in row:
                if cell.value is not None:
                    label_list.append(cell.value)
        else:
            for j, cell in enumerate(row):
                if j == 0:
                    xlabel_list.append(cell.value)
                elif j == 1:
                    if cell.value is None:
                        origin_list.append(0.0)
                    else:
                        origin_list.append(float(cell.value))
                elif j == 2:
                    if cell.value is None:
                        pretrained_list.append(0.0)
                    else:
                        pretrained_list.append(float(cell.value))
                elif j == 3:
                    if cell.value is None:
                        spp_list.append(0.0)
                    else:
                        spp_list.append(float(cell.value))
    print("不同的条件有：", label_list)
    
    print("实验模型有：", xlabel_list)
    
    print("origin条件下结果：", origin_list)
    print("pretrained条件下结果：", pretrained_list)
    print("spp条件下结果：", spp_list)
    return label_list, xlabel_list, origin_list, pretrained_list, spp_list

if __name__ == "__main__":
    label_list, xlabel_list, origin_list, pretrained_list, spp_list = get_excel_data(excel_path='all_result.xlsx')
    data_excel = pd.DataFrame(origin_list, columns=['origin'], index=xlabel_list)
    data_excel = pd.concat([data_excel, pd.DataFrame(pretrained_list, columns=['pretrained'], index=xlabel_list)], axis=1)
    data_excel = pd.concat([data_excel, pd.DataFrame(spp_list, columns=['spp'], index=xlabel_list)], axis=1)
    print(data_excel)
    data_excel.plot(kind='bar', style='--o', figsize=(20, 15), title="")
    plt.show()